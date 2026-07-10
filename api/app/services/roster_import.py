"""Roster xlsx importer (SPRD §5.6 / P0-C) — students mode.

Ported analyze→confirm→commit flow + keyword-heuristic column mapping from the fee
system (the AI step is optional and stubbed; the heuristic is what runs without an
AI key and is tuned to the school's real registers). Stateless: analyze returns
the parsed rows, the client confirms the mapping, and commit re-sends both.
"""

import io
import uuid
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.context import CurrentMember
from app.models import Guardian, SchoolClass, Student, StudentCategory
from app.services.ai.extract import phrase_gap_question

# Unified student roster target fields the importer can populate.
TARGET_FIELDS = [
    "full_name", "admission_no", "roll_no", "class_name", "section", "category",
    "father_name", "father_phone", "mother_name", "mother_phone", "phone",
]

# Without these two, commit() rejects the row (no name / no admission no), so they
# are the only gaps worth stopping the admin for. The rest degrade quietly.
REQUIRED_FIELDS = ("full_name", "admission_no")

# Shown in the gap question when a required column can't be found.
FIELD_LABELS: dict[str, str] = {
    "full_name": "name", "admission_no": "admission number", "roll_no": "roll number",
    "class_name": "class", "section": "section", "category": "category",
}

FIELD_HINTS: dict[str, list[str]] = {
    "full_name": ["student name", "name", "student", "full name", "child name"],
    "admission_no": ["sr no", "sr. no", "adm", "admission", "adm. no", "scholar", "enrollment"],
    "roll_no": ["roll number", "roll no", "roll"],
    "class_name": ["class", "grade", "standard", "std", "class name"],
    "section": ["section", "sec", "div", "division"],
    "category": ["category", "student category", "type"],
    "father_name": ["father's name", "father name", "father"],
    "father_phone": ["father's mobile", "father mobile", "father's phone", "father phone"],
    "mother_name": ["mother's name", "mother name", "mother"],
    "mother_phone": ["mother's mobile", "mother mobile", "mother's phone", "mother phone"],
    "phone": ["contact number", "contact", "mobile", "phone", "mobile no", "guardian"],
}


def read_first_sheet(data: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    """Parse the first non-empty sheet into (columns, rows). The header is the
    first row with any non-empty cell."""
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header_idx = next(
                (i for i, r in enumerate(rows) if any(c is not None and str(c).strip() for c in r)),
                None,
            )
            if header_idx is None:
                continue
            header = [
                str(c).strip() if c is not None else f"col_{i}"
                for i, c in enumerate(rows[header_idx])
            ]
            records: list[dict[str, Any]] = []
            for r in rows[header_idx + 1:]:
                if not any(c is not None and str(c).strip() for c in r):
                    continue
                records.append({
                    h: (str(r[i]).strip() if i < len(r) and r[i] is not None else None)
                    for i, h in enumerate(header)
                })
            return [h for h in header if h], records
    finally:
        wb.close()
    return [], []


def heuristic_mapping(columns: list[str]) -> dict[str, str]:
    """Keyword-match source columns to student fields (heuristic fallback)."""
    lc = {c.lower().strip(): c for c in columns}
    mapping: dict[str, str] = {}
    for field in TARGET_FIELDS:
        for hint in FIELD_HINTS.get(field, []):
            hit = next((orig for low, orig in lc.items() if hint == low or hint in low), None)
            if hit:
                mapping[field] = hit
                break
    return mapping


def analyze(data: bytes) -> dict[str, Any]:
    columns, rows = read_first_sheet(data)
    mapping = heuristic_mapping(columns)
    # Same envelope as the staff and syllabus importers (see services/ingest.py), so
    # the one import panel can render all three. The mapping stays roster's own tuned
    # heuristic; only the surrounding gap fields are added. A required column we could
    # not place becomes a tap-to-answer question rather than a wall of per-row errors
    # at commit — and their absence used to crash the panel, which read `.length` on
    # fields this endpoint never returned.
    used = set(mapping.values())
    unmapped = [c for c in columns if c not in used]
    missing = [f for f in REQUIRED_FIELDS if f not in mapping]
    questions = [
        phrase_gap_question("students", f, FIELD_LABELS.get(f, f), unmapped) for f in missing
    ]
    # Return ALL rows (not just a preview): the flow is stateless, so the client
    # holds the parsed rows and sends them back to /commit with the confirmed
    # mapping. Fine for roster sizes (hundreds); no server-side file cache.
    return {
        "columns": columns,
        "mapping": mapping,
        "rows": rows,
        "row_count": len(rows),
        "unmapped_columns": unmapped,
        "missing_required": missing,
        "low_confidence": [],
        "questions": questions,
        "source": "heuristic",
    }


class RosterImporter:
    def __init__(self, db: Session):
        self.db = db

    def _class_map(self, org_id: uuid.UUID, year_id: uuid.UUID | None) -> dict[tuple, uuid.UUID]:
        if year_id is None:
            return {}
        rows = self.db.scalars(
            select(SchoolClass).where(
                SchoolClass.org_id == org_id, SchoolClass.academic_year_id == year_id
            )
        )
        return {(c.name.strip().lower(), (c.section or "").strip().lower()): c.id for c in rows}

    def _get_or_create_category(self, org_id: uuid.UUID, name: str,
                                cache: dict[str, uuid.UUID]) -> uuid.UUID:
        key = name.strip().lower()
        if key in cache:
            return cache[key]
        existing = self.db.scalar(
            select(StudentCategory).where(
                StudentCategory.org_id == org_id, StudentCategory.name.ilike(name.strip())
            )
        )
        if existing is None:
            existing = StudentCategory(org_id=org_id, name=name.strip())
            self.db.add(existing)
            self.db.flush()
        cache[key] = existing.id
        return existing.id

    def commit(self, m: CurrentMember, *, mapping: dict[str, str], rows: list[dict],
               academic_year_id: uuid.UUID | None) -> dict[str, Any]:
        classes = self._class_map(m.org_id, academic_year_id)
        cat_cache: dict[str, uuid.UUID] = {}
        # Existing admission numbers to skip duplicates in one query.
        existing_adm = set(self.db.scalars(
            select(Student.admission_no).where(Student.org_id == m.org_id)
        ))
        created = 0
        skipped = 0
        errors: list[dict[str, Any]] = []

        def val(row: dict, field: str) -> str | None:
            col = mapping.get(field)
            v = row.get(col) if col else None
            return v.strip() if isinstance(v, str) and v.strip() else None

        for idx, row in enumerate(rows):
            name = val(row, "full_name")
            adm = val(row, "admission_no")
            if not name:
                errors.append({"row": idx + 1, "reason": "missing name"})
                continue
            if not adm:
                errors.append({"row": idx + 1, "reason": "missing admission no."})
                continue
            if adm in existing_adm:
                skipped += 1
                continue

            class_id = None
            cname, section = val(row, "class_name"), val(row, "section")
            if cname:
                class_id = classes.get((cname.lower(), (section or "").lower()))

            category_id = None
            cat = val(row, "category")
            if cat:
                category_id = self._get_or_create_category(m.org_id, cat, cat_cache)

            student = Student(
                org_id=m.org_id, admission_no=adm, full_name=name,
                roll_no=val(row, "roll_no"), class_id=class_id, category_id=category_id,
            )
            self.db.add(student)
            self.db.flush()
            existing_adm.add(adm)

            # Guardians: father (primary), mother, or a lone phone.
            for gname_f, gphone_f, relation, primary in [
                ("father_name", "father_phone", "Father", True),
                ("mother_name", "mother_phone", "Mother", False),
            ]:
                gname, gphone = val(row, gname_f), val(row, gphone_f)
                if gname or gphone:
                    self.db.add(Guardian(
                        org_id=m.org_id, student_id=student.id, name=gname or relation,
                        relation=relation, phone=gphone or "", is_primary=primary,
                    ))
            lone = val(row, "phone")
            if lone and not val(row, "father_phone") and not val(row, "mother_phone"):
                self.db.add(Guardian(
                    org_id=m.org_id, student_id=student.id, name="Guardian",
                    phone=lone, is_primary=True,
                ))
            created += 1

        self.db.flush()
        return {"created": created, "skipped": skipped, "errors": errors}
